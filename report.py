"""
實價登錄月報產生器（Excel）
Sheet 1: 總覽摘要
Sheet 2: 各區月度成交排名
Sheet 3: 建案明細
"""

import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

DB_PATH = "real_estate.db"
OUTPUT_PATH = "實價登錄月報.xlsx"

C_DARK   = "1A2E44"
C_MID    = "2E6DA4"
C_LIGHT  = "D6E8F7"
C_ACCENT = "E8571A"
C_GOLD   = "F4B942"
C_WHITE  = "FFFFFF"
C_GRAY   = "F5F5F5"
C_BORDER = "CCCCCC"
FONT     = "Arial"


def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")


def bd(color=C_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def cell_style(cell, value=None, bold=False, size=10, color="000000",
               bg=None, align="left", wrap=False, num_fmt=None):
    if value is not None:
        cell.value = value
    cell.font = Font(name=FONT, bold=bold, size=size, color=color)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = bd()
    if num_fmt:
        cell.number_format = num_fmt


def header(ws, row, col, text, bg=C_DARK, size=11, span=1):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name=FONT, bold=True, size=size, color=C_WHITE)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = bd()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)


def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ── 載入資料 ─────────────────────────────────────────────

def _roc_to_ym(val):
    try:
        if hasattr(val, 'year'):
            return f"{val.year}-{val.month:02d}"
        s = str(val).strip().replace("/", "").replace(".", "").replace("-", "")
        s = s.split(".")[0].zfill(7)
        return f"{int(s[:3]) + 1911}-{s[3:5]}"
    except Exception:
        return ""


def load_data():
    conn = sqlite3.connect(DB_PATH)
    try:
        df = pd.read_sql("SELECT * FROM presale", conn)
    except Exception:
        df = pd.DataFrame()
    try:
        log_df = pd.read_sql(
            "SELECT 年月, 資料來源, 新增筆數, 匯入時間 FROM month_log ORDER BY 年月 DESC",
            conn)
    except Exception:
        log_df = pd.DataFrame()
    conn.close()

    if df.empty:
        return df, log_df

    df["總價元"] = pd.to_numeric(df["總價元"], errors="coerce")
    df["單價元平方公尺"] = pd.to_numeric(df["單價元平方公尺"], errors="coerce")
    df["建物移轉總面積平方公尺"] = pd.to_numeric(
        df["建物移轉總面積平方公尺"], errors="coerce")
    df["總價萬"] = (df["總價元"] / 10000).round(1)
    df["單價萬坪"] = (df["單價元平方公尺"] * 3.3058 / 10000).round(2)
    df["面積坪"] = (df["建物移轉總面積平方公尺"] * 0.3025).round(1)

    if "年月" not in df.columns or df["年月"].isna().all():
        df["年月"] = df["交易年月日"].apply(_roc_to_ym)

    return df, log_df


# ── Sheet 1：總覽摘要 ─────────────────────────────────────

def make_summary(wb, df, log_df):
    ws = wb.active
    ws.title = "總覽摘要"
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 24

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "大台中預售屋實價登錄月報"
    c.font = Font(name=FONT, bold=True, size=18, color=C_WHITE)
    c.fill = PatternFill("solid", start_color=C_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    months_str = "、".join(log_df["年月"].tolist()) if not log_df.empty else "-"
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    c.value = f"資料月份：{months_str}　　更新時間：{now}　　資料來源：內政部實價登錄"
    c.font = Font(name=FONT, size=10, color=C_WHITE)
    c.fill = PatternFill("solid", start_color=C_MID)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # KPI
    row = 4
    ws.merge_cells(f"A{row}:H{row}")
    header(ws, row, 1, "▌ 本期整體概況", bg=C_MID, size=12, span=8)
    ws.row_dimensions[row].height = 28

    row = 5
    kpi_headers = ["月份數", "累積筆數", "均單價（萬/坪）",
                   "中位單價（萬/坪）", "均總價（萬）",
                   "最高總價（萬）", "最低總價（萬）", "有建案名稱（筆）"]
    for i, h in enumerate(kpi_headers, 1):
        cell_style(ws.cell(row=row, column=i), h,
                   bold=True, bg=C_LIGHT, align="center", size=10)
    ws.row_dimensions[row].height = 22

    row = 6
    kpi_vals = [
        len(log_df),
        len(df),
        round(df["單價萬坪"].mean(), 2) if not df.empty else 0,
        round(df["單價萬坪"].median(), 2) if not df.empty else 0,
        round(df["總價萬"].mean(), 1) if not df.empty else 0,
        round(df["總價萬"].max(), 1) if not df.empty else 0,
        round(df["總價萬"].min(), 1) if not df.empty else 0,
        int((df["建案名稱"].fillna("").str.strip() != "").sum()) if not df.empty else 0,
    ]
    for i, v in enumerate(kpi_vals, 1):
        c = ws.cell(row=row, column=i)
        cell_style(c, v, bold=True, size=12, align="center", bg=C_WHITE)
    ws.row_dimensions[row].height = 30

    # 各月統計
    row = 8
    ws.merge_cells(f"A{row}:H{row}")
    header(ws, row, 1, "▌ 各月成交統計", bg=C_MID, size=12, span=8)
    ws.row_dimensions[row].height = 28

    row = 9
    m_headers = ["成交年月", "筆數", "均單（萬/坪）", "中位單（萬/坪）",
                 "均總（萬）", "最高單（萬/坪）", "最低單（萬/坪）", "建案名稱有值"]
    for i, h in enumerate(m_headers, 1):
        cell_style(ws.cell(row=row, column=i), h,
                   bold=True, bg=C_LIGHT, align="center")
    ws.row_dimensions[row].height = 22

    month_grp = pd.DataFrame()
    if not df.empty and "年月" in df.columns:
        month_grp = df.groupby("年月").agg(
            筆數=("id", "count"),
            均單=("單價萬坪", "mean"),
            中位單=("單價萬坪", "median"),
            均總=("總價萬", "mean"),
            最高單=("單價萬坪", "max"),
            最低單=("單價萬坪", "min"),
        ).reset_index().sort_values("年月", ascending=False)

        for ri, (_, r) in enumerate(month_grp.iterrows()):
            row = 10 + ri
            bg = C_GRAY if ri % 2 == 0 else C_WHITE
            has_name = int((df[df["年月"] == r["年月"]]["建案名稱"].fillna("").str.strip() != "").sum())
            vals = [r["年月"], int(r["筆數"]),
                    round(r["均單"], 2), round(r["中位單"], 2),
                    round(r["均總"], 1), round(r["最高單"], 2),
                    round(r["最低單"], 2), has_name]
            for i, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=i)
                cell_style(c, v, bg=bg, align="center" if i > 1 else "left")

    # 各行政區統計
    row = 10 + len(month_grp) + 2
    ws.merge_cells(f"A{row}:H{row}")
    header(ws, row, 1, "▌ 各行政區累積概況", bg=C_MID, size=12, span=8)
    ws.row_dimensions[row].height = 28

    row += 1
    d_headers = ["行政區", "筆數", "均單（萬/坪）", "中位單（萬/坪）",
                 "均總（萬）", "最高總（萬）", "最低總（萬）", "建案數"]
    for i, h in enumerate(d_headers, 1):
        cell_style(ws.cell(row=row, column=i), h,
                   bold=True, bg=C_LIGHT, align="center")
    ws.row_dimensions[row].height = 22

    if not df.empty:
        dist_grp = df.groupby("鄉鎮市區").agg(
            筆數=("id", "count"),
            均單=("單價萬坪", "mean"),
            中位單=("單價萬坪", "median"),
            均總=("總價萬", "mean"),
            最高總=("總價萬", "max"),
            最低總=("總價萬", "min"),
            建案數=("建案名稱", lambda x: x[x.str.strip() != ""].nunique()),
        ).reset_index().sort_values("筆數", ascending=False)

        for ri, (_, r) in enumerate(dist_grp.iterrows()):
            row += 1
            bg = C_GRAY if ri % 2 == 0 else C_WHITE
            vals = [r["鄉鎮市區"], int(r["筆數"]),
                    round(r["均單"], 2), round(r["中位單"], 2),
                    round(r["均總"], 1), round(r["最高總"], 1),
                    round(r["最低總"], 1), int(r["建案數"])]
            for i, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=i)
                cell_style(c, v, bg=bg, align="center" if i > 1 else "left")

    set_col_widths(ws, {
        "A": 14, "B": 10, "C": 16, "D": 16,
        "E": 14, "F": 14, "G": 14, "H": 12
    })
    ws.freeze_panes = "A3"
    log("Sheet 1 總覽摘要 完成")


# ── Sheet 2：各區月度成交排名 ────────────────────────────

def make_district_ranking(wb, df):
    ws = wb.create_sheet("各區月度排名")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = "各行政區月度成交排名"
    c.font = Font(name=FONT, bold=True, size=16, color=C_WHITE)
    c.fill = PatternFill("solid", start_color=C_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")

    if df.empty or "年月" not in df.columns:
        log("Sheet 2 無資料")
        return

    months = sorted(df["年月"].unique(), reverse=True)
    row = 3

    for ym in months:
        mdf = df[df["年月"] == ym]

        ws.merge_cells(f"A{row}:K{row}")
        header(ws, row, 1, f"▌ {ym}", bg=C_MID, size=12, span=11)
        ws.row_dimensions[row].height = 26
        row += 1

        cols = ["排名", "行政區", "成交筆數", "均單價（萬/坪）",
                "中位單（萬/坪）", "最高單（萬/坪）", "最低單（萬/坪）",
                "均總價（萬）", "最高總（萬）", "最低總（萬）", "建案數"]
        for i, h in enumerate(cols, 1):
            cell_style(ws.cell(row=row, column=i), h,
                       bold=True, bg=C_LIGHT, align="center", size=10)
        ws.row_dimensions[row].height = 22
        row += 1

        dist = mdf.groupby("鄉鎮市區").agg(
            筆數=("id", "count"),
            均單=("單價萬坪", "mean"),
            中位單=("單價萬坪", "median"),
            最高單=("單價萬坪", "max"),
            最低單=("單價萬坪", "min"),
            均總=("總價萬", "mean"),
            最高總=("總價萬", "max"),
            最低總=("總價萬", "min"),
            建案數=("建案名稱", lambda x: x[x.str.strip() != ""].nunique()),
        ).reset_index().sort_values("筆數", ascending=False).reset_index(drop=True)

        for ri, (_, r) in enumerate(dist.iterrows()):
            bg = C_GOLD if ri == 0 else (C_GRAY if ri % 2 == 0 else C_WHITE)
            vals = [ri + 1, r["鄉鎮市區"], int(r["筆數"]),
                    round(r["均單"], 2), round(r["中位單"], 2),
                    round(r["最高單"], 2), round(r["最低單"], 2),
                    round(r["均總"], 1), round(r["最高總"], 1),
                    round(r["最低總"], 1), int(r["建案數"])]
            for i, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=i)
                cell_style(c, v, bold=(ri == 0), bg=bg,
                           align="center" if i != 2 else "left")
            row += 1

        row += 1

    set_col_widths(ws, {
        "A": 6, "B": 12, "C": 10, "D": 16, "E": 16,
        "F": 16, "G": 16, "H": 14, "I": 12, "J": 12, "K": 8
    })
    ws.freeze_panes = "A3"
    log("Sheet 2 各區月度排名 完成")


# ── Sheet 3：建案明細 ─────────────────────────────────────

def make_case_detail(wb, df):
    ws = wb.create_sheet("建案明細")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = "建案成交明細（依月份 × 行政區）"
    c.font = Font(name=FONT, bold=True, size=16, color=C_WHITE)
    c.fill = PatternFill("solid", start_color=C_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")

    if df.empty or "年月" not in df.columns:
        log("Sheet 3 無資料")
        return

    has_name = df[df["建案名稱"].fillna("").str.strip() != ""].copy()
    if has_name.empty:
        ws.cell(row=3, column=1, value="無建案名稱資料")
        log("Sheet 3 無建案名稱資料")
        return

    months = sorted(df["年月"].unique(), reverse=True)
    row = 3

    for ym in months:
        mdf = has_name[has_name["年月"] == ym]
        if mdf.empty:
            continue

        ws.merge_cells(f"A{row}:K{row}")
        header(ws, row, 1, f"▌ {ym}", bg=C_MID, size=12, span=11)
        ws.row_dimensions[row].height = 26
        row += 1

        for dist in sorted(mdf["鄉鎮市區"].unique()):
            ddf = mdf[mdf["鄉鎮市區"] == dist]

            ws.merge_cells(f"A{row}:K{row}")
            header(ws, row, 1, f"  {dist}", bg="4A7FB5", size=10, span=11)
            ws.row_dimensions[row].height = 20
            row += 1

            cols = ["建案名稱", "行政區", "成交筆數",
                    "均單（萬/坪）", "中位單（萬/坪）",
                    "最高單（萬/坪）", "最低單（萬/坪）",
                    "均總（萬）", "最高總（萬）", "最低總（萬）", "建物型態"]
            for i, h in enumerate(cols, 1):
                cell_style(ws.cell(row=row, column=i), h,
                           bold=True, bg=C_LIGHT, align="center", size=9)
            ws.row_dimensions[row].height = 20
            row += 1

            cases = ddf.groupby("建案名稱").agg(
                行政區=("鄉鎮市區", "first"),
                筆數=("id", "count"),
                均單=("單價萬坪", "mean"),
                中位單=("單價萬坪", "median"),
                最高單=("單價萬坪", "max"),
                最低單=("單價萬坪", "min"),
                均總=("總價萬", "mean"),
                最高總=("總價萬", "max"),
                最低總=("總價萬", "min"),
                建物型態=("建物型態", lambda x: x.mode()[0] if len(x) > 0 else ""),
            ).reset_index().sort_values("筆數", ascending=False)

            for ri, (_, r) in enumerate(cases.iterrows()):
                bg = C_GOLD if ri == 0 else (C_GRAY if ri % 2 == 0 else C_WHITE)
                vals = [r["建案名稱"], r["行政區"], int(r["筆數"]),
                        round(r["均單"], 2), round(r["中位單"], 2),
                        round(r["最高單"], 2), round(r["最低單"], 2),
                        round(r["均總"], 1), round(r["最高總"], 1),
                        round(r["最低總"], 1), r["建物型態"]]
                for i, v in enumerate(vals, 1):
                    c = ws.cell(row=row, column=i)
                    cell_style(c, v, bold=(ri == 0), bg=bg,
                               align="left" if i in [1, 2, 11] else "center",
                               size=9)
                row += 1

            row += 1
        row += 1

    set_col_widths(ws, {
        "A": 20, "B": 10, "C": 8, "D": 14, "E": 14,
        "F": 14, "G": 14, "H": 12, "I": 12, "J": 12, "K": 20
    })
    ws.freeze_panes = "A3"
    log("Sheet 3 建案明細 完成")


# ── 主程式 ────────────────────────────────────────────────

def generate_report():
    log("=== 開始產生月報 ===")
    df, log_df = load_data()

    if df.empty:
        log("資料庫無資料，終止")
        return

    log(f"載入 {len(df):,} 筆，{len(log_df)} 個月")

    wb = Workbook()
    make_summary(wb, df, log_df)
    make_district_ranking(wb, df)
    make_case_detail(wb, df)

    wb.save(OUTPUT_PATH)
    log(f"=== 月報已儲存：{OUTPUT_PATH} ===")


if __name__ == "__main__":
    generate_report()
